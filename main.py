from __future__ import annotations

import argparse
import csv
import gzip
import io
import hashlib
import json
import os
import re
import shutil
import subprocess
import tempfile
import time
import unicodedata
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import quote, unquote, urlparse, urlsplit, urlunsplit

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
import requests
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ============================================================
# BI INSS - UNIVERSO PREVIDENCIARIO
# INSS 2025+ -> Parquet -> SharePoint -> Power BI
# ============================================================

APP_VERSION = "5.3.0"

CKAN_BASE = "https://dadosabertos.inss.gov.br"
CKAN_ORG = "instituto-nacional-de-seguro-social-inss"
ANO_INICIAL = int(os.getenv("ANO_INICIAL", "2025"))

SITE_ID = os.getenv(
    "SHAREPOINT_SITE_ID",
    "universoprevidenciario.sharepoint.com,"
    "b4ca13e8-7517-4e58-bb82-4cb8f8e06866,"
    "c2bf221a-26c3-4ff9-8dd3-648cb5cc1552",
)
BASE_FOLDER = os.getenv("SHAREPOINT_BASE_FOLDER", "INSS")

CSV_CHUNK_ROWS = 100_000
DOWNLOAD_CHUNK = 8 * 1024 * 1024
UPLOAD_CHUNK = 10 * 1024 * 1024   # multiplo de 320 KiB
SIMPLE_UPLOAD_LIMIT = 240 * 1024 * 1024

RULES = {
    "concedidos": "beneficios concedidos",
    "emitidos": "beneficios emitidos",
    "indeferidos": "beneficios indeferidos",
    "mantidos": "beneficios mantidos",
    "cat": "comunicacoes de acidente de trabalho",
    "solicitados": "dados quantitativos de requerimentos administrativos solicitados",
    "pendentes": "dados quantitativos de requerimentos administrativos pendentes",
    "unidades": "perfil das unidades",
    "folha_emitidos": "dados agregados da folha de pagamento em relacao aos beneficios emitidos",
}


class SourceUnavailable(RuntimeError):
    """Fonte oficial temporariamente inacessivel (ex.: 403/404)."""



def norm(value: Any) -> str:
    s = "" if value is None else str(value)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s.lower().strip())


def safe_name(value: Any, limit: int = 100) -> str:
    s = re.sub(r"[^a-z0-9._-]+", "-", norm(value))
    return (re.sub(r"-+", "-", s).strip("-._") or "arquivo")[:limit]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def years_of(resource: dict[str, Any]) -> list[int]:
    # O ano deve vir do proprio recurso, nunca do titulo/faixa do dataset.
    # Ordem: nome do recurso -> nome do arquivo na URL -> descricao.
    candidates = [
        str(resource.get("name") or ""),
        Path(urlparse(str(resource.get("url") or "")).path).name,
        str(resource.get("description") or ""),
    ]
    for text in candidates:
        years = sorted({int(x) for x in re.findall(r"\b(20\d{2})\b", text)})
        if years:
            return years
    return []


def resource_is_valid(resource: dict[str, Any]) -> bool:
    years = years_of(resource)
    return bool(years) and max(years) >= ANO_INICIAL


def resource_version(resource: dict[str, Any]) -> str:
    return str(
        resource.get("last_modified")
        or resource.get("created")
        or resource.get("url")
        or resource.get("id")
        or ""
    )


def build_session() -> requests.Session:
    retry = Retry(
        total=5,
        backoff_factor=1,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET", "HEAD", "POST", "PUT", "DELETE"}),
        respect_retry_after_header=True,
    )
    s = requests.Session()
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.headers.update({"User-Agent": "UniversoPrevidenciario-BI-INSS/5.1.0"})
    return s


HTTP = build_session()


def ckan(action: str, params: dict[str, Any]) -> Any:
    r = HTTP.get(
        f"{CKAN_BASE}/api/3/action/{action}",
        params=params,
        timeout=(30, 180),
    )
    r.raise_for_status()
    payload = r.json()
    if not payload.get("success"):
        raise RuntimeError(f"CKAN success=false: {action}")
    return payload["result"]


def load_catalog() -> list[dict[str, Any]]:
    result = ckan(
        "package_search",
        {"q": f"organization:{CKAN_ORG}", "rows": 1000, "start": 0},
    )
    return list(result.get("results", []))


def dataset_matches(category: str, title: str) -> bool:
    t = norm(title)
    rule = RULES[category]

    if category in {"concedidos", "emitidos", "indeferidos", "mantidos", "cat"}:
        return t.startswith(rule)

    return rule in t


def choose_dataset(
    packages: list[dict[str, Any]],
    category: str,
) -> dict[str, Any]:
    candidates = [
        p for p in packages
        if dataset_matches(category, str(p.get("title") or ""))
    ]
    if not candidates:
        raise RuntimeError(f"Dataset nao encontrado: {category}")

    def score(p: dict[str, Any]) -> tuple[int, int, str]:
        resources = p.get("resources") or []
        valid_count = sum(resource_is_valid(r) for r in resources)
        current_plan = int("plano de dados abertos" in norm(p.get("title")))
        return valid_count, current_plan, str(p.get("metadata_modified") or "")

    return max(candidates, key=score)


def list_resources(dataset: dict[str, Any]) -> list[dict[str, Any]]:
    resources = [r for r in (dataset.get("resources") or []) if resource_is_valid(r)]
    resources.sort(
        key=lambda r: (
            max(years_of(r) or [0]),
            str(r.get("last_modified") or r.get("created") or ""),
            str(r.get("name") or ""),
        ),
        reverse=True,
    )
    return resources


class Graph:
    def __init__(self) -> None:
        self.session = build_session()
        self.token = ""
        self.drive_id = ""
        self.root_id = ""
        self.refresh_token()
        self.load_drive()

    def refresh_token(self) -> None:
        result = subprocess.run(
            [
                "az", "account", "get-access-token",
                "--resource-type", "ms-graph",
                "--query", "accessToken",
                "--output", "tsv",
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        self.token = result.stdout.strip()
        if not self.token:
            raise RuntimeError("Token Microsoft Graph vazio.")

    def request(
        self,
        method: str,
        endpoint: str,
        *,
        json_body: Any | None = None,
        data: Any | None = None,
        headers: dict[str, str] | None = None,
        expected: tuple[int, ...] = (200,),
        retry_auth: bool = True,
    ) -> requests.Response:
        url = endpoint if endpoint.startswith("http") else f"https://graph.microsoft.com/v1.0{endpoint}"
        h = {"Authorization": f"Bearer {self.token}"}
        if headers:
            h.update(headers)

        r = self.session.request(
            method,
            url,
            json=json_body,
            data=data,
            headers=h,
            timeout=(30, 300),
        )

        if r.status_code == 401 and retry_auth:
            self.refresh_token()
            return self.request(
                method,
                endpoint,
                json_body=json_body,
                data=data,
                headers=headers,
                expected=expected,
                retry_auth=False,
            )

        if r.status_code not in expected:
            raise RuntimeError(f"Graph HTTP {r.status_code}: {r.text[:1200]}")
        return r

    def load_drive(self) -> None:
        self.drive_id = self.request(
            "GET",
            f"/sites/{SITE_ID}/drive?$select=id",
        ).json()["id"]
        self.root_id = self.request(
            "GET",
            f"/drives/{self.drive_id}/root?$select=id",
        ).json()["id"]

    def get_item_by_path(self, path: str) -> requests.Response:
        enc = quote(path.strip("/"), safe="/")
        url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{enc}"
        r = self.session.get(
            url,
            headers={"Authorization": f"Bearer {self.token}"},
            timeout=(30, 120),
        )
        if r.status_code == 401:
            self.refresh_token()
            r = self.session.get(
                url,
                headers={"Authorization": f"Bearer {self.token}"},
                timeout=(30, 120),
            )
        return r

    def ensure_folder(self, path: str) -> None:
        parts = [p for p in path.strip("/").split("/") if p]
        parent_id = self.root_id
        current: list[str] = []

        for part in parts:
            current.append(part)
            current_path = "/".join(current)
            r = self.get_item_by_path(current_path)

            if r.status_code == 200:
                parent_id = r.json()["id"]
                continue
            if r.status_code != 404:
                raise RuntimeError(f"Falha pasta {current_path}: HTTP {r.status_code}")

            try:
                created = self.request(
                    "POST",
                    f"/drives/{self.drive_id}/items/{parent_id}/children",
                    json_body={
                        "name": part,
                        "folder": {},
                        "@microsoft.graph.conflictBehavior": "fail",
                    },
                    expected=(201,),
                )
                parent_id = created.json()["id"]
            except RuntimeError:
                r2 = self.get_item_by_path(current_path)
                if r2.status_code != 200:
                    raise
                parent_id = r2.json()["id"]

    def read_json(self, path: str) -> dict[str, Any]:
        enc = quote(path.strip("/"), safe="/")
        url = f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}/root:/{enc}:/content"
        r = self.session.get(
            url,
            headers={"Authorization": f"Bearer {self.token}"},
            timeout=(30, 120),
            allow_redirects=True,
        )
        if r.status_code == 404:
            return {}
        if r.status_code == 401:
            self.refresh_token()
            return self.read_json(path)
        if r.status_code != 200:
            raise RuntimeError(f"Falha ao ler {path}: HTTP {r.status_code}")
        return r.json()

    def upload_bytes(self, content: bytes, remote: str, content_type: str) -> None:
        folder = str(Path(remote).parent).replace("\\", "/")
        if folder != ".":
            self.ensure_folder(folder)

        enc = quote(remote.strip("/"), safe="/")
        self.request(
            "PUT",
            f"/drives/{self.drive_id}/root:/{enc}:/content",
            data=content,
            headers={"Content-Type": content_type},
            expected=(200, 201),
        )

    def upload_file(self, local: Path, remote: str) -> None:
        folder = str(Path(remote).parent).replace("\\", "/")
        if folder != ".":
            self.ensure_folder(folder)

        if local.stat().st_size <= SIMPLE_UPLOAD_LIMIT:
            enc = quote(remote.strip("/"), safe="/")
            url = (
                f"https://graph.microsoft.com/v1.0/drives/{self.drive_id}"
                f"/root:/{enc}:/content"
            )

            for attempt in range(1, 6):
                with local.open("rb") as f:
                    r = requests.put(
                        url,
                        data=f,
                        headers={
                            "Authorization": f"Bearer {self.token}",
                            "Content-Type": "application/octet-stream",
                        },
                        timeout=(30, 300),
                    )

                if r.status_code in (200, 201):
                    return
                if r.status_code == 401:
                    self.refresh_token()
                    continue
                if r.status_code in (429, 500, 502, 503, 504):
                    time.sleep(min(30, attempt * 3))
                    continue
                raise RuntimeError(
                    f"Upload simples HTTP {r.status_code}: {r.text[:800]}"
                )

            raise RuntimeError("Upload simples excedeu tentativas.")

        enc = quote(remote.strip("/"), safe="/")
        session_info = self.request(
            "POST",
            f"/drives/{self.drive_id}/root:/{enc}:/createUploadSession",
            json_body={
                "item": {
                    "@microsoft.graph.conflictBehavior": "replace",
                    "name": Path(remote).name,
                }
            },
        ).json()

        upload_url = session_info["uploadUrl"]
        total = local.stat().st_size

        with local.open("rb") as f:
            start = 0
            while start < total:
                chunk = f.read(UPLOAD_CHUNK)
                if not chunk:
                    break

                end = start + len(chunk) - 1
                headers = {
                    "Content-Length": str(len(chunk)),
                    "Content-Range": f"bytes {start}-{end}/{total}",
                }

                for attempt in range(1, 6):
                    r = requests.put(
                        upload_url,
                        data=chunk,
                        headers=headers,
                        timeout=(30, 300),
                    )
                    if r.status_code in (200, 201, 202):
                        break
                    if r.status_code in (429, 500, 502, 503, 504):
                        time.sleep(min(30, attempt * 3))
                        continue
                    raise RuntimeError(f"Upload grande HTTP {r.status_code}: {r.text[:800]}")
                else:
                    raise RuntimeError("Upload grande excedeu tentativas.")

                start = end + 1


def detect_csv_bytes(sample: bytes) -> tuple[str, str, int]:
    """Detecta encoding, delimitador e linha real do cabecalho."""
    encoding = "utf-8"
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            sample.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            pass

    text = sample.decode(encoding)
    lines = text.splitlines()

    for idx, line in enumerate(lines[:10]):
        stripped = line.strip()
        if not stripped:
            continue
        m = re.fullmatch(r"sep=(.)", stripped, flags=re.I)
        if m:
            return encoding, m.group(1), idx + 1
        break

    def width(line: str, delimiter: str) -> int:
        try:
            return len(
                next(
                    csv.reader(
                        [line],
                        delimiter=delimiter,
                        quotechar='"',
                        strict=False,
                    )
                )
            )
        except (csv.Error, StopIteration):
            return 0

    candidates = [";", ",", "\t", "|", "^"]
    useful = [
        line for line in lines[:120]
        if line.strip() and not line.lstrip().startswith(("#", "//"))
    ]

    best = None
    for delimiter in candidates:
        widths = [width(line, delimiter) for line in useful[:80]]
        widths = [w for w in widths if w > 1]
        if not widths:
            continue

        counts: dict[int, int] = {}
        for w in widths:
            counts[w] = counts.get(w, 0) + 1

        mode_width, mode_count = max(
            counts.items(),
            key=lambda x: (x[1], x[0]),
        )
        consistency = mode_count / len(widths)
        score = (consistency, mode_count, mode_width)

        if best is None or score > best[0]:
            best = (score, delimiter, mode_width)

    if best is None:
        raise RuntimeError("Nao foi possivel detectar o delimitador do CSV.")

    _, delimiter, dominant_width = best

    header_index = 0
    for i, line in enumerate(lines[:50]):
        if not line.strip():
            continue
        if width(line, delimiter) != dominant_width:
            continue
        try:
            fields = next(
                csv.reader(
                    [line],
                    delimiter=delimiter,
                    quotechar='"',
                    strict=False,
                )
            )
        except csv.Error:
            continue

        nonempty = sum(bool(str(v).strip()) for v in fields)
        if nonempty >= max(2, dominant_width // 2):
            header_index = i
            break

    return encoding, delimiter, header_index


def detect_csv(path: Path) -> tuple[str, str, int]:
    with path.open("rb") as f:
        sample = f.read(512_000)
    return detect_csv_bytes(sample)


def unique_columns(columns: list[Any]) -> list[str]:
    used: dict[str, int] = {}
    result: list[str] = []

    for i, c in enumerate(columns, start=1):
        base = str(c).strip().lstrip("\ufeff") if c is not None else ""
        base = base or f"coluna_{i}"
        used[base] = used.get(base, 0) + 1
        result.append(base if used[base] == 1 else f"{base}__{used[base]}")
    return result


def chunk_to_table(df: pd.DataFrame) -> pa.Table:
    columns = unique_columns(list(df.columns))
    df.columns = columns

    arrays = {}
    for c in columns:
        arrays[c] = pa.array(
            [None if pd.isna(v) else str(v) for v in df[c].tolist()],
            type=pa.string(),
        )
    return pa.table(arrays)


def _csv_factory_to_parquet(
    factory,
    target: Path,
    *,
    encoding: str,
    delimiter: str,
    header_index: int,
    label: str,
) -> int:
    """Converte sem descartar linhas; falha se a estrutura continuar inconsistente."""
    standard_attempts = [
        {
            "name": "csv_padrao",
            "quoting": csv.QUOTE_MINIMAL,
            "doublequote": True,
        },
        {
            "name": "aspas_com_escape",
            "quoting": csv.QUOTE_MINIMAL,
            "doublequote": False,
            "escapechar": "\\",
        },
        {
            "name": "aspas_literais",
            "quoting": csv.QUOTE_NONE,
        },
    ]

    # Os arquivos Mantidos de jul/2026 publicados pelo INSS possuem aspas
    # inconsistentes. Para eles, tratar aspas como texto literal evita uma
    # leitura completa de dezenas de GB antes de cair no fallback.
    if "mantidos" in norm(label):
        attempts = [standard_attempts[2], standard_attempts[0], standard_attempts[1]]
    else:
        attempts = standard_attempts

    last_error: Exception | None = None

    for attempt_number, options in enumerate(attempts, start=1):
        target.unlink(missing_ok=True)
        writer: pq.ParquetWriter | None = None
        rows = 0

        try:
            with factory() as handle:
                reader = pd.read_csv(
                    handle,
                    sep=delimiter,
                    encoding=encoding,
                    skiprows=header_index,
                    engine="python",
                    dtype=str,
                    chunksize=CSV_CHUNK_ROWS,
                    keep_default_na=False,
                    na_filter=False,
                    on_bad_lines="error",
                    quoting=options["quoting"],
                    doublequote=options.get("doublequote", True),
                    escapechar=options.get("escapechar"),
                )

                for chunk in reader:
                    table = chunk_to_table(chunk)
                    if writer is None:
                        writer = pq.ParquetWriter(
                            target,
                            table.schema,
                            compression="zstd",
                            use_dictionary=True,
                        )
                    writer.write_table(table)
                    rows += len(chunk)

            if writer is None:
                raise RuntimeError(f"CSV sem dados: {label}")

            writer.close()
            writer = None
            print(
                f"CSV OK | {label} | modo={options['name']} | "
                f"{rows:,} linhas"
            )
            return rows

        except (pd.errors.ParserError, csv.Error, UnicodeDecodeError) as exc:
            last_error = exc
            print(
                f"AVISO | parser {attempt_number}/{len(attempts)} falhou "
                f"em {label}: {type(exc).__name__}: {exc}"
            )
        finally:
            if writer is not None:
                try:
                    writer.close()
                except Exception:
                    pass

        target.unlink(missing_ok=True)

    raise RuntimeError(
        f"CSV oficial continua inconsistente apos {len(attempts)} "
        f"estrategias: {label}. Ultimo erro: {last_error}"
    )


def csv_to_parquet(source: Path, target: Path) -> int:
    encoding, delimiter, header_index = detect_csv(source)
    print(
        f"CSV detectado | encoding={encoding} | "
        f"delimitador={delimiter!r} | cabecalho={header_index + 1}"
    )
    return _csv_factory_to_parquet(
        lambda: source.open("rb"),
        target,
        encoding=encoding,
        delimiter=delimiter,
        header_index=header_index,
        label=source.name,
    )


def zip_csv_to_parquet(
    archive: zipfile.ZipFile,
    member: str,
    target: Path,
) -> int:
    with archive.open(member) as handle:
        sample = handle.read(512_000)

    encoding, delimiter, header_index = detect_csv_bytes(sample)
    print(
        f"ZIP/CSV detectado | {member} | encoding={encoding} | "
        f"delimitador={delimiter!r} | cabecalho={header_index + 1}"
    )

    return _csv_factory_to_parquet(
        lambda: archive.open(member),
        target,
        encoding=encoding,
        delimiter=delimiter,
        header_index=header_index,
        label=member,
    )


def xlsx_to_parquet(source: Path, output_dir: Path, prefix: str) -> list[tuple[Path, int]]:
    wb = load_workbook(source, read_only=True, data_only=True)
    outputs: list[tuple[Path, int]] = []

    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            header = None

            for _ in range(30):
                try:
                    row = next(it)
                except StopIteration:
                    break
                if any(v is not None and str(v).strip() for v in row):
                    header = unique_columns(list(row))
                    break

            if not header:
                continue

            suffix = f"__{safe_name(ws.title)}" if len(wb.worksheets) > 1 else ""
            target = output_dir / f"{prefix}{suffix}.parquet"
            schema = pa.schema([(c, pa.string()) for c in header])
            writer = pq.ParquetWriter(target, schema, compression="zstd")
            rows = 0
            batch: list[list[Any]] = []

            try:
                for row in it:
                    batch.append(list(row))
                    if len(batch) >= 50_000:
                        data = {
                            c: pa.array(
                                [
                                    None if i >= len(r) or r[i] is None else str(r[i])
                                    for r in batch
                                ],
                                type=pa.string(),
                            )
                            for i, c in enumerate(header)
                        }
                        writer.write_table(pa.table(data))
                        rows += len(batch)
                        batch.clear()

                if batch:
                    data = {
                        c: pa.array(
                            [
                                None if i >= len(r) or r[i] is None else str(r[i])
                                for r in batch
                            ],
                            type=pa.string(),
                        )
                        for i, c in enumerate(header)
                    }
                    writer.write_table(pa.table(data))
                    rows += len(batch)
            finally:
                writer.close()

            outputs.append((target, rows))
    finally:
        wb.close()

    if not outputs:
        raise RuntimeError(f"XLSX sem dados tabulares: {source.name}")
    return outputs



def xls_to_parquet(source: Path, output_dir: Path, prefix: str) -> list[tuple[Path, int]]:
    sheets = pd.read_excel(source, sheet_name=None, dtype=str, engine="xlrd")
    outputs: list[tuple[Path, int]] = []

    for sheet_name, df in sheets.items():
        if df.empty and len(df.columns) == 0:
            continue
        df.columns = unique_columns(list(df.columns))
        suffix = f"__{safe_name(sheet_name)}" if len(sheets) > 1 else ""
        target = output_dir / f"{prefix}{suffix}.parquet"
        table = chunk_to_table(df)
        pq.write_table(table, target, compression="zstd", use_dictionary=True)
        outputs.append((target, len(df)))

    if not outputs:
        raise RuntimeError(f"XLS sem dados tabulares: {source.name}")
    return outputs

def detect_file_kind(path: Path) -> str:
    """Detecta o formato REAL pelo conteudo, nao pela etiqueta do CKAN."""
    with path.open("rb") as f:
        head = f.read(16)

    if head.startswith(b"\x1f\x8b"):
        return "gzip"

    if zipfile.is_zipfile(path):
        # XLSX tambem e um ZIP, por isso diferenciamos pelo conteudo interno.
        try:
            with zipfile.ZipFile(path) as z:
                names = {n.replace("\\", "/") for n in z.namelist()}
                if "[Content_Types].xml" in names and any(n.startswith("xl/") for n in names):
                    return "xlsx"
        except zipfile.BadZipFile:
            pass
        return "zip"

    if head.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "xls"

    stripped = head.lstrip().lower()
    if stripped.startswith(b"<html") or stripped.startswith(b"<!doctype"):
        return "html"

    return "csv"


def _download_url_variants(url: str) -> list[str]:
    """Tenta tambem a variante em que '+' publicado no caminho representa espaco."""
    variants = [url]
    parsed = urlsplit(url)
    decoded_path = unquote(parsed.path)

    if "+" in decoded_path:
        alternate_path = quote(decoded_path.replace("+", " "), safe="/")
        alternate = urlunsplit(
            (
                parsed.scheme,
                parsed.netloc,
                alternate_path,
                parsed.query,
                parsed.fragment,
            )
        )
        if alternate not in variants:
            variants.append(alternate)

    return variants


def _download_via_datastore(
    resource: dict[str, Any],
    folder: Path,
) -> tuple[Path, str, str] | None:
    """Fallback CKAN DataStore quando o arquivo externo estiver indisponivel."""
    rid = str(resource.get("id") or "")
    if not rid:
        return None

    endpoint = f"{CKAN_BASE}/api/3/action/datastore_search"

    try:
        probe = HTTP.get(
            endpoint,
            params={"resource_id": rid, "limit": 1, "offset": 0},
            timeout=(30, 180),
        )
        if probe.status_code != 200:
            return None

        payload = probe.json()
        if not payload.get("success"):
            return None

        result = payload.get("result") or {}
        fields = [
            str(f.get("id"))
            for f in (result.get("fields") or [])
            if str(f.get("id") or "") != "_id"
        ]
        if not fields:
            return None

        total = int(result.get("total") or 0)
        target = folder / "origem_datastore.csv"

        print(
            f"Fallback CKAN DataStore ativo | recurso={rid} | "
            f"registros={total:,}"
        )

        page_size = 10_000
        offset = 0

        with target.open("w", encoding="utf-8", newline="") as out:
            writer = csv.DictWriter(
                out,
                fieldnames=fields,
                delimiter=";",
                quoting=csv.QUOTE_MINIMAL,
                extrasaction="ignore",
                lineterminator="\n",
            )
            writer.writeheader()

            while offset < total:
                response = HTTP.get(
                    endpoint,
                    params={
                        "resource_id": rid,
                        "limit": page_size,
                        "offset": offset,
                    },
                    timeout=(30, 300),
                )
                response.raise_for_status()
                page = response.json()
                if not page.get("success"):
                    raise RuntimeError(
                        f"CKAN DataStore success=false no offset {offset}"
                    )

                records = (page.get("result") or {}).get("records") or []
                if not records:
                    break

                for record in records:
                    clean = {}
                    for field in fields:
                        value = record.get(field)
                        if value is None:
                            clean[field] = ""
                        elif isinstance(value, (dict, list)):
                            clean[field] = json.dumps(
                                value,
                                ensure_ascii=False,
                                separators=(",", ":"),
                            )
                        else:
                            clean[field] = str(value)
                    writer.writerow(clean)

                offset += len(records)
                print(
                    f"DataStore | {min(offset, total):,}/{total:,} registros"
                )

        digest = hashlib.sha256()
        with target.open("rb") as handle:
            for chunk in iter(lambda: handle.read(DOWNLOAD_CHUNK), b""):
                digest.update(chunk)

        return target, digest.hexdigest(), "csv"

    except (requests.RequestException, ValueError, KeyError, TypeError) as exc:
        print(f"DataStore indisponivel para {rid}: {exc}")
        return None


def download_resource(resource: dict[str, Any], folder: Path) -> tuple[Path, str, str]:
    url = str(resource.get("url") or "")
    if not url:
        raise RuntimeError("Recurso sem URL.")

    statuses: list[str] = []

    for candidate_url in _download_url_variants(url):
        raw = folder / "origem.bin"
        raw.unlink(missing_ok=True)
        digest = hashlib.sha256()

        try:
            with HTTP.get(
                candidate_url,
                stream=True,
                headers={
                    "Referer": f"{CKAN_BASE}/",
                    "Accept": "*/*",
                },
                timeout=(30, 600),
                allow_redirects=True,
            ) as r:
                if r.status_code in (403, 404):
                    statuses.append(f"{r.status_code} {candidate_url}")
                    print(
                        f"Fonte externa respondeu HTTP {r.status_code}; "
                        "tentando fallback..."
                    )
                    continue

                r.raise_for_status()
                content_type = str(r.headers.get("Content-Type") or "")

                with raw.open("wb") as f:
                    for chunk in r.iter_content(DOWNLOAD_CHUNK):
                        if chunk:
                            f.write(chunk)
                            digest.update(chunk)

            if raw.stat().st_size == 0:
                statuses.append(f"arquivo vazio {candidate_url}")
                continue

            kind = detect_file_kind(raw)
            if kind == "html":
                statuses.append(
                    f"HTML em vez de dados {candidate_url} "
                    f"Content-Type={content_type!r}"
                )
                raw.unlink(missing_ok=True)
                continue

            suffix = {
                "zip": ".zip",
                "gzip": ".gz",
                "xlsx": ".xlsx",
                "xls": ".xls",
                "csv": ".csv",
            }[kind]
            target = folder / f"origem{suffix}"
            target.unlink(missing_ok=True)
            raw.replace(target)
            return target, digest.hexdigest(), kind

        except requests.RequestException as exc:
            statuses.append(f"{type(exc).__name__}: {exc}")
            raw.unlink(missing_ok=True)

    datastore = _download_via_datastore(resource, folder)
    if datastore is not None:
        return datastore

    raise SourceUnavailable(
        "Fonte oficial indisponivel e sem fallback DataStore. "
        + " | ".join(statuses[-3:])
    )


def convert_resource(
    source: Path,
    resource: dict[str, Any],
    folder: Path,
) -> list[tuple[Path, int]]:
    out = folder / "out"
    out.mkdir(exist_ok=True)

    prefix = f"{safe_name(resource.get('id'), 50)}__{safe_name(resource.get('name'), 80)}"
    kind = detect_file_kind(source)

    if kind == "csv":
        target = out / f"{prefix}.parquet"
        return [(target, csv_to_parquet(source, target))]

    if kind == "xlsx":
        return xlsx_to_parquet(source, out, prefix)

    if kind == "xls":
        return xls_to_parquet(source, out, prefix)

    if kind == "gzip":
        extracted = folder / "gzip_extraido.bin"
        with gzip.open(source, "rb") as src_gz, extracted.open("wb") as dst:
            shutil.copyfileobj(src_gz, dst, length=DOWNLOAD_CHUNK)
        inner_kind = detect_file_kind(extracted)
        suffix = {
            "zip": ".zip",
            "xlsx": ".xlsx",
            "xls": ".xls",
            "csv": ".csv",
        }.get(inner_kind)
        if not suffix:
            raise RuntimeError(f"GZIP contem formato nao suportado: {inner_kind}")
        renamed = folder / f"gzip_extraido{suffix}"
        extracted.replace(renamed)
        return convert_resource(renamed, resource, folder)

    if kind != "zip":
        raise RuntimeError(f"Formato real nao suportado: {kind}")

    with zipfile.ZipFile(source) as z:
        members = []
        for n in z.namelist():
            if n.endswith("/") or "__MACOSX" in n:
                continue
            parts = Path(n.replace("\\", "/")).parts
            if ".." in parts:
                raise RuntimeError(f"ZIP com caminho inseguro: {n}")
            members.append(n)

        csvs = [n for n in members if Path(n).suffix.lower() in {".csv", ".txt"}]
        excels = [n for n in members if Path(n).suffix.lower() in {".xlsx", ".xls"}]

        # Se o ZIP trouxer CSV + JSON espelho, usamos somente o tabular.
        chosen = csvs or excels
        if not chosen:
            raise RuntimeError("ZIP sem CSV/TXT/XLSX/XLS tabular.")

        outputs: list[tuple[Path, int]] = []

        for i, member in enumerate(chosen, start=1):
            declared_ext = Path(member).suffix.lower()
            inner_prefix = f"{prefix}__{safe_name(Path(member).stem)}"

            # CSV/TXT grande e lido diretamente do ZIP.
            if declared_ext in {".csv", ".txt"}:
                info = z.getinfo(member)
                print(
                    f"ZIP/CSV streaming | {member} | "
                    f"{info.file_size / 1024 / 1024:,.1f} MB descompactado"
                )
                target = out / f"{inner_prefix}.parquet"
                outputs.append(
                    (target, zip_csv_to_parquet(z, member, target))
                )
                continue

            extracted_raw = folder / f"zip_{i}.bin"
            with z.open(member) as src_zip, extracted_raw.open("wb") as dst:
                shutil.copyfileobj(src_zip, dst, length=DOWNLOAD_CHUNK)

            inner_kind = detect_file_kind(extracted_raw)
            actual_ext = {
                "xlsx": ".xlsx",
                "xls": ".xls",
                "zip": ".zip",
                "csv": ".csv",
            }.get(inner_kind, declared_ext)
            extracted = folder / f"zip_{i}{actual_ext}"
            extracted_raw.replace(extracted)

            try:
                if inner_kind == "xlsx":
                    outputs.extend(
                        xlsx_to_parquet(extracted, out, inner_prefix)
                    )
                elif inner_kind == "xls":
                    outputs.extend(
                        xls_to_parquet(extracted, out, inner_prefix)
                    )
                elif inner_kind == "csv":
                    target = out / f"{inner_prefix}.parquet"
                    outputs.append(
                        (target, csv_to_parquet(extracted, target))
                    )
                elif inner_kind == "zip":
                    nested_resource = dict(resource)
                    nested_resource["name"] = (
                        f"{resource.get('name')} - {Path(member).stem}"
                    )
                    outputs.extend(
                        convert_resource(
                            extracted,
                            nested_resource,
                            folder,
                        )
                    )
                else:
                    raise RuntimeError(
                        f"Formato interno nao suportado: {inner_kind}"
                    )
            finally:
                extracted.unlink(missing_ok=True)

        return outputs

def control_path(category: str) -> str:
    return f"{BASE_FOLDER}/_controle/{category}.json"


def load_control(graph: Graph, category: str) -> dict[str, Any]:
    control = graph.read_json(control_path(category))
    if not control:
        control = {}
    if not isinstance(control, dict):
        raise RuntimeError(f"Controle invalido para {category}")
    resources = control.setdefault("resources", {})
    if not isinstance(resources, dict):
        raise RuntimeError(f"Controle.resources invalido para {category}")
    return control


def save_control(graph: Graph, category: str, control: dict[str, Any]) -> None:
    control["updated_at_utc"] = now_iso()
    graph.upload_bytes(
        json.dumps(control, ensure_ascii=False, indent=2).encode("utf-8"),
        control_path(category),
        "application/json; charset=utf-8",
    )


def record_blocked_source(
    graph: Graph,
    control: dict[str, Any],
    category: str,
    resource: dict[str, Any],
    error: Exception,
) -> None:
    rid = str(resource.get("id") or resource.get("url") or safe_name(resource.get("name")))
    blocked = control.setdefault("blocked_sources", {})
    previous = blocked.get(rid) if isinstance(blocked, dict) else None
    if not isinstance(blocked, dict):
        blocked = {}
        control["blocked_sources"] = blocked
    attempts = int((previous or {}).get("attempts") or 0) + 1
    blocked[rid] = {
        "status": "source_unavailable",
        "category": category,
        "resource_name": resource.get("name"),
        "resource_url": resource.get("url"),
        "version": resource_version(resource),
        "attempts": attempts,
        "last_seen_utc": now_iso(),
        "error": str(error)[:2000],
    }
    save_control(graph, category, control)


def needs_processing(
    control: dict[str, Any],
    resource: dict[str, Any],
    force: bool,
) -> bool:
    if force:
        return True

    rid = str(resource.get("id") or "")
    previous = control["resources"].get(rid)
    return not previous or previous.get("version") != resource_version(resource)


def process_one(
    graph: Graph,
    control: dict[str, Any],
    category: str,
    dataset: dict[str, Any],
    resource: dict[str, Any],
) -> None:
    rid = str(resource.get("id") or "")
    name = str(resource.get("name") or rid)
    year = str(max(years_of(resource)))

    print("\n" + "=" * 80)
    print(f"{category.upper()} | {name}")
    print("=" * 80)

    with tempfile.TemporaryDirectory(prefix="bi_inss_") as tmp:
        folder = Path(tmp)

        print("1/4 Download oficial...")
        source, sha, detected_kind = download_resource(resource, folder)
        print(f"OK | {source.stat().st_size / 1024 / 1024:,.1f} MB")
        print(f"Formato declarado: {resource.get('format')} | mimetype={resource.get('mimetype')}")
        print(f"Formato REAL detectado: {detected_kind.upper()}")

        print("2/4 Conversao Parquet em lotes...")
        outputs = convert_resource(source, resource, folder)
        print(f"OK | {sum(r for _, r in outputs):,} linhas")

        print("3/4 Upload SharePoint...")
        remote_outputs = []
        for local, rows in outputs:
            remote = f"{BASE_FOLDER}/silver/{category}/{year}/{local.name}"
            graph.upload_file(local, remote)
            remote_outputs.append(
                {"path": remote, "rows": rows, "bytes": local.stat().st_size}
            )
            print(f"OK | {remote}")

        print("4/4 Checkpoint...")
        control["resources"][rid] = {
            "category": category,
            "dataset_id": dataset.get("id"),
            "dataset_slug": dataset.get("name"),
            "dataset_title": dataset.get("title"),
            "resource_name": name,
            "resource_url": resource.get("url"),
            "resource_format": resource.get("format"),
            "resource_mimetype": resource.get("mimetype"),
            "detected_kind": detected_kind,
            "app_version": APP_VERSION,
            "version": resource_version(resource),
            "source_sha256": sha,
            "years": years_of(resource),
            "processed_at_utc": now_iso(),
            "outputs": remote_outputs,
        }
        blocked = control.get("blocked_sources")
        if isinstance(blocked, dict):
            blocked.pop(rid, None)
        save_control(graph, category, control)
        print("CHECKPOINT OK")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--dataset",
        choices=["all", *RULES.keys()],
        default=os.getenv("INSS_DATASET", "all"),
    )
    parser.add_argument(
        "--max-resources",
        type=int,
        default=int(os.getenv("MAX_RESOURCES", "3")),
        help="0 = sem limite",
    )
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--catalog-only", action="store_true")
    args = parser.parse_args()

    categories = list(RULES) if args.dataset == "all" else [args.dataset]

    print("=" * 80)
    print("UNIVERSO PREVIDENCIARIO - BI INSS CLOUD")
    print(f"APP_VERSION: {APP_VERSION}")
    print(f"Periodo: {ANO_INICIAL} em diante")
    print("Processamento: um recurso por vez")
    print("=" * 80)

    packages = load_catalog()
    print(f"Datasets encontrados no portal: {len(packages)}")

    selected: dict[str, dict[str, Any]] = {}
    for category in categories:
        ds = choose_dataset(packages, category)
        selected[category] = ds
        print(
            f"[{category}] {ds.get('title')} | "
            f"{len(list_resources(ds))} recursos {ANO_INICIAL}+"
        )

    if args.catalog_only:
        print("CATALOGO OK - nenhum arquivo baixado.")
        return 0

    graph = Graph()
    graph.ensure_folder(f"{BASE_FOLDER}/_controle")
    print("SharePoint conectado.")

    ok = 0
    errors = 0
    blocked_sources = 0

    for category, dataset in selected.items():
        control = load_control(graph, category)
        pending = [
            r for r in list_resources(dataset)
            if needs_processing(control, r, args.force)
        ]

        target_successes = (
            len(pending)
            if args.max_resources <= 0
            else min(args.max_resources, len(pending))
        )

        print(
            f"\n[{category}] pendentes totais: {len(pending)} | "
            f"meta de sucessos nesta execucao: {target_successes}"
        )

        successes_in_category = 0

        for resource in pending:
            if (
                args.max_resources > 0
                and successes_in_category >= args.max_resources
            ):
                break

            try:
                process_one(
                    graph,
                    control,
                    category,
                    dataset,
                    resource,
                )
                ok += 1
                successes_in_category += 1

            except SourceUnavailable as e:
                blocked_sources += 1
                record_blocked_source(
                    graph, control, category, resource, e
                )
                print(
                    f"BLOQUEIO NA FONTE OFICIAL | {category} | "
                    f"{resource.get('name')}"
                )
                print(str(e))
                print(
                    f"::warning title=Fonte oficial indisponivel::{category} | "
                    f"{resource.get('name')} sera tentado novamente na proxima execucao."
                )
                print(
                    "Continuando para o proximo recurso sem marcar "
                    "este como concluido."
                )

            except Exception as e:
                errors += 1
                print(f"ERRO | {category} | {resource.get('name')}")
                print(str(e))
                print(
                    "Nao foi marcado como concluido; "
                    "sera tentado novamente."
                )

    print("\n" + "=" * 80)
    print(f"SUCESSO: {ok}")
    print(f"ERROS DE PROCESSAMENTO: {errors}")
    print(f"FONTES OFICIAIS BLOQUEADAS: {blocked_sources}")
    print("=" * 80)

    # Erro de processamento deixa o job vermelho.
    # Fonte oficial temporariamente bloqueada vira WARNING auditavel e sera
    # retentada nas proximas execucoes, sem derrubar toda a carga disponivel.
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
